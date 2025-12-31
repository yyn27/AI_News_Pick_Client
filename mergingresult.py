# 合并一个月的所有数据文件，并修正 URL 列为真实超链接
#!/usr/bin/env python
# -*- coding: utf-8 -*-

import pandas as pd
from pathlib import Path
import re
import openpyxl

# ====== 你需要改这里 ======
FOLDER = r"D:\원일남\데이터크롤링결과\12월"   # 改成你的文件夹路径
PATTERN = "*.xlsx"                # 匹配的文件类型
SHEET_NAME = 0                    # 第一个工作表；如要指定可填表名字符串
OUTPUT = r"D:\원일남\데이터크롤링결과\12월\merge_result-all.xlsx"      # 输出文件名
TARGET_COLS = ["원본기사"]         # 需要强制写回 URL 的列名（可加其他列）
# ==========================

HYPERLINK_RE = re.compile(r'HYPERLINK\s*\(\s*"([^"]+)"\s*(?:,|\))', re.IGNORECASE)

def url_from_cell(cell):
    """优先取超链接对象；否则解析 HYPERLINK 公式；否则返回 None"""
    if cell.hyperlink and getattr(cell.hyperlink, "target", None):
        return cell.hyperlink.target
    v = cell.value
    if isinstance(v, str) and v.startswith("="):
        m = HYPERLINK_RE.search(v.lstrip("="))
        if m:
            return m.group(1)
    return None

def overwrite_urls_from_wb(xlsx_path: Path, df: pd.DataFrame) -> pd.DataFrame:
    """用 openpyxl 直接读取工作簿，把 TARGET_COLS 中的列整列覆盖为真实 URL"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=False, read_only=False)
    ws = wb[wb.sheetnames[0] if isinstance(SHEET_NAME, int) else SHEET_NAME]

    # 取表头 → 列名到列号
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    name2col = {str(v): i+1 for i, v in enumerate(header) if v is not None}

    for col_name in TARGET_COLS:
        if col_name not in df.columns or col_name not in name2col:
            continue
        cidx = name2col[col_name]
        urls = []
        # 只遍历 df 的行数，保证对齐：df 第 0 行 ↔ ws 第 2 行
        for i in range(len(df)):
            cell = ws.cell(row=i+2, column=cidx)
            url = url_from_cell(cell)
            # 如果取不到 URL，就保留 df 里原值
            urls.append(url if url else df.iloc[i][col_name])
        # 整列覆盖
        df[col_name] = urls

    wb.close()
    return df

def read_one(p: Path) -> pd.DataFrame:
    # 用 pandas 读其余内容（dtype=object 保持原样）
    df = pd.read_excel(p, sheet_name=SHEET_NAME, engine="openpyxl", dtype=object)

    # 关键：强制覆盖 URL 列
    df = overwrite_urls_from_wb(p, df)

    # 删除每个文件最后 4 行
    if len(df) >= 4:
        df = df.iloc[:-4, :].copy()

    # 溯源列
    df["source_file"] = p.name
    return df

def main():
    files = sorted(Path(FOLDER).glob(PATTERN))
    if not files:
        raise SystemExit("未找到待合并的文件")

    parts = []
    for f in files:
        try:
            parts.append(read_one(f))
            print(f"[OK] {f.name}")
        except Exception as e:
            print(f"[ERR] {f.name}: {e}")

    merged = pd.concat(parts, ignore_index=True)
    merged.to_excel(OUTPUT, index=False)
    print(f"✅ DONE: {OUTPUT} (rows={len(merged)})")

if __name__ == "__main__":
    main()