import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
import openpyxl
import os
import time

# 设置文件路径
output_path = r"D:\AI_News_Pick_Client\naver_blog_kity4099_all_250821.xlsx"

# 如果文件不存在，创建一个新文件（含表头）
if not os.path.exists(output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Naver Blog"
    ws.append(["게시글URL", "게시글제목", "등록일자", "계정명", "게시글내용"])
    wb.save(output_path)

# 加载 Excel
wb = openpyxl.load_workbook(output_path)
ws = wb.active

# 已经爬取过的链接集合（防止重复）
visited_links = set(ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1))

# 创建两个 driver
options_scroll = uc.ChromeOptions()
options_scroll.add_argument('--headless')
options_scroll.add_argument('--disable-gpu')
options_scroll.add_argument("--lang=ko-KR")

options_detail = uc.ChromeOptions()
options_detail.add_argument('--headless')
options_detail.add_argument('--disable-gpu')
options_detail.add_argument("--lang=ko-KR")

driver_scroll = uc.Chrome(options=options_scroll)
driver_detail = uc.Chrome(options=options_detail)

# 打开博客主页
blog_id = 'kity4099'
url = f"https://m.blog.naver.com/PostList.naver?blogId={blog_id}&categoryNo=0&listStyle=post&tab=1"
driver_scroll.get(url)
time.sleep(3)

SCROLL_PAUSE_TIME = 2
no_change_count = 0
max_no_change = 4
last_count = 0
total_processed = len(visited_links)

print("🚀 开始滚动并实时抓取文章...")

while True:
    # 下拉到底部加载更多
    driver_scroll.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(SCROLL_PAUSE_TIME)

    # 提取当前页面中的所有文章链接
    soup = BeautifulSoup(driver_scroll.page_source, 'html.parser')
    new_tags = soup.select("a.link__A4O1D")
    all_links = [tag['href'] for tag in new_tags if tag.has_attr('href')]

    # 去重后逐条爬取新链接
    for post_url in all_links:
        if post_url in visited_links:
            continue

        print(f"🔎 正在抓取第 {total_processed+1} 篇: {post_url}")
        try:
            driver_detail.get(post_url)
            time.sleep(2)
            post_soup = BeautifulSoup(driver_detail.page_source, 'html.parser')

            # 提取标题
            title_tag = post_soup.select_one('.se-title-text') or post_soup.find('h3')
            title = title_tag.get_text(strip=True) if title_tag else "无标题"

            # 发布时间
            date_tag = post_soup.select_one('p.blog_date')
            pub_date = date_tag.get_text(strip=True) if date_tag else "未知"

            # 계정명（作者昵称）
            author_tag = post_soup.select_one('strong.ell')
            author_name = author_tag.get_text(strip=True) if author_tag else "알 수 없음"  # "无法识别"

            # 正文内容
            content_tag = post_soup.select_one('.se-main-container')
            content = content_tag.get_text("\n", strip=True) if content_tag else "未找到正文"

            # 写入 Excel（实时保存）
            ws.append([post_url, title, pub_date, author_name, content])
            wb.save(output_path)
            visited_links.add(post_url)
            total_processed += 1

        except Exception as e:
            print(f"❌ 出错: {e}")
            continue

    # 停止条件判断：若连续几次未加载新链接，认为到底
    current_count = len(all_links)
    if current_count == last_count:
        no_change_count += 1
        if no_change_count >= max_no_change:
            print("✅ 已加载全部页面，程序结束！")
            break
    else:
        no_change_count = 0
        last_count = current_count

# 退出浏览器
driver_scroll.quit()
driver_detail.quit()
print(f"\n✅ 爬取完成，总共写入 {total_processed} 篇文章。文件已保存为：{output_path}")
