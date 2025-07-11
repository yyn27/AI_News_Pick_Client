# ✅ 修改后的 core/preprocessing.py

import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import logging

import sys
def resource_path(relative_path):
    """兼容PyInstaller和源码运行的资源路径"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

INPUT_PATH = os.environ.get("INPUT_EXCEL_PATH", resource_path("data/input/default.xlsx"))
OUTPUT_PATH = os.environ.get("OUTPUT_EXCEL_PATH", resource_path(f"data/output/output_{datetime.now().strftime('%y%m%d')}.xlsx"))

# ==== 로그 설정 ====
log_dir = resource_path("data/log")
os.makedirs(log_dir, exist_ok=True)
today = datetime.now().strftime("%y%m%d")
log_file_path = os.path.join(log_dir, f"로그_{today}.txt")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file_path, encoding='utf-8'),
        logging.StreamHandler()
    ]
)

def log(msg):
    logging.info(msg)

def preprocess_title(title):
    return title.split('&keyword=')[0] if isinstance(title, str) else title

def read_excel_with_hyperlinks(file_path, sheet_name=0):
    wb = load_workbook(filename=file_path, data_only=True, read_only=False)
    sheet = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]
    data = []
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    for row in sheet.iter_rows(min_row=2):
        row_data = {}
        for col_idx, cell in enumerate(row):
            header = headers[col_idx]
            value = cell.value
            hyperlink = cell.hyperlink.target if cell.hyperlink else None
            if header == "게시글제목":
                row_data[header] = value
                row_data["게시글URL"] = hyperlink
            else:
                row_data[header] = value
        data.append(row_data)
    return pd.DataFrame(data)

def filter_untrusted_posts(all_data):
    untrusted_file = resource_path("resources/비신탁사_저작권문구+도메인주소.xlsx")
    trusted_file = resource_path("resources/매체사_도메인_정보.xlsx")

    df_untrusted = pd.read_excel(untrusted_file)
    df_trusted = pd.read_excel(trusted_file)
    untrusted_copyrights = df_untrusted["저작권 문구"].dropna().tolist()
    untrusted_domains = df_untrusted["도메인"].dropna().tolist()
    trusted_domains = df_trusted["도메인"].dropna().tolist()

    def should_remove(post_content):
        post_content = str(post_content)
        contains_untrusted_copyright = any(c in post_content for c in untrusted_copyrights)
        contains_untrusted_domain = any(d in post_content for d in untrusted_domains)
        contains_trusted_domain = any(d in post_content for d in trusted_domains)
        return (contains_untrusted_copyright or contains_untrusted_domain) and not contains_trusted_domain

    mask = all_data["게시글내용"].apply(should_remove)
    df_filtered = all_data[~mask]
    df_removed = all_data[mask]

    log(f"비신탁사 필터링 완료: 유지 {len(df_filtered)}개 / 삭제 {len(df_removed)}개")
    return df_filtered, df_removed

def filter_empty_image_and_no_da(df_filtered):
    mask = (
        ((~df_filtered["게시글제목"].str.contains("다.", regex=False, na=False)) |
         (df_filtered["게시글제목"].str.contains("니다.", regex=False, na=False))) &
        ((~df_filtered["게시글내용"].str.contains("다.", regex=False, na=False)) |
         (df_filtered["게시글내용"].str.contains("니다.", regex=False, na=False))) &
        (~df_filtered["게시글제목"].str.contains("만평", regex=False, na=False)) &
        (~df_filtered["게시글내용"].str.contains("만평", regex=False, na=False))
    )
    df_final = df_filtered[~mask]
    df_removed_images = df_filtered[mask]

    log(f"텍스트 필터링 완료: 유지 {len(df_final)}개 / 삭제 {len(df_removed_images)}개")
    return df_final, df_removed_images

def run_preprocessing(input_path=None, output_path=None, stop_event=None):
    all_data = read_excel_with_hyperlinks(input_path)
    all_data.columns = [str(col).strip() for col in all_data.columns]
    all_data['게시글제목'] = all_data['게시글제목'].apply(preprocess_title)

    exclude_file_path = resource_path("resources/(언진) 수집 제외 도메인 주소_공식 블로그-0709.xlsx")
    exclude_df = pd.read_excel(exclude_file_path)
    exclude_urls = exclude_df['제외 도메인 주소(블로그)'].dropna().astype(str).tolist()
    filtered_data = all_data[~all_data['게시글URL'].astype(str).apply(
        lambda url: any(excluded in url for excluded in exclude_urls)
    )]

    log(f"총 행 수: {len(all_data)}")
    log(f"제외된 후 남은 행 수: {len(filtered_data)}")

    all_df_drop_search = filtered_data[
        (filtered_data.apply(lambda x: x['검색어'].lower() in str(x['게시글제목']).lower() or
                                      x['검색어'].lower() in str(x['게시글내용']).lower(), axis=1)) &
        (~filtered_data['게시글내용'].fillna('').str.contains('신춘문예', case=False)) &
        (~filtered_data['게시글제목'].fillna('').str.contains('신춘문예', case=False)) &
        (~filtered_data['계정명'].fillna('').str.contains('뽐뿌뉴스', case=False))
    ]
    log(f"삭제 : {len(filtered_data) - len(all_df_drop_search)}개")
    log(all_df_drop_search.count())

    if all_df_drop_search.empty:
        log("⚠️ 검색어 기반 필터링 결과: 남은 행이 없습니다. 전처리를 중단합니다.")
        all_df_drop_search.head(0).to_excel(output_path, index=False)
        log(f"✅ 전처리 완료. 저장됨 → {output_path}")
        return
    if stop_event and stop_event.is_set():
        log("🛑 사용자 중단 요청 감지, 작업 중단")
        return

    df_filtered, _ = filter_untrusted_posts(all_df_drop_search)
    if df_filtered.empty:
        log("⚠️ 비신탁사 필터링 결과: 남은 행이 없습니다. 전처리를 중단합니다.")
        df_filtered.head(0).to_excel(output_path, index=False)
        log(f"✅ 전처리 완료. 저장됨 → {output_path}")
        return
    if stop_event and stop_event.is_set():
        log("🛑 사용자 중단 요청 감지, 작업 중단")
        return
    
    df_final, _ = filter_empty_image_and_no_da(df_filtered)
    if df_final.empty:
        log("⚠️ 텍스트 필터링 결과: 남은 행이 없습니다. 전처리를 중단합니다.")
        df_final.head(0).to_excel(output_path, index=False)
        log(f"✅ 전처리 완료. 저장됨 → {output_path}")
        return
    if stop_event and stop_event.is_set():
        log("🛑 사용자 중단 요청 감지, 작업 중단")
        return

    df_final.to_excel(output_path, index=False)
    log(f"✅ 전처리 완료. 저장됨 → {output_path}")

if __name__ == "__main__":
    run_preprocessing()
